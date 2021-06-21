import pandas as pd
import string
import random
from random import choice
import xlwt
from xlwt import Workbook
import openpyxl
  
# Workbook is created
#wb = Workbook()
#wbkName = 'D:/vt/Create-populate-excel-questions.xlsx'
#wbk = openpyxl.load_workbook(wbkName)
#sheet1 = wbk.worksheets
#sheet1 = None
#if 'sheet1' in wbk.sheetnames:
 #   sheet1 = wb['sheet1']
#sheet1.write(1, 0, 'ISBT DEHRADUN')
#sheet1.write(2, 0, 'SHASTRADHARA')
#wb.save('xlwt example.xls')  
# add_sheet is used to create sheet.
#sheet1 = wb.add_sheet('Sheet 1')
df = pd.read_excel("D:/vt/Create-populate-excel-questions.xlsx", "Sheet1")
#print(df.to_dict('r'))
#print(df["questionId"])
df["Statement"]
df["Answer"]
lst = df.to_dict("r")
lst1 = list()
lst2 = list()

def string_for_year_to_days(y, d):
    str1 =  "Convert {y} year {d} days to days (year is a leap year).".format(y=y, d=d)
   # res = (y*365)+d
    return str1

def conver_year_to_days(y, d):
    str1 =  "Convert {y} year {d} days to days (year is a leap year).".format(y=y, d=d)
    res = (y*365)+d
    return res

def string_year_and_sum_months_to_months(y, m):
    str1 =  "Convert {y} years {m} months into months.".format(y=y, m=m)
   # res = (y*365)+d
    return str1

def conver_year_sum_monts_to_months(y, m):
    res = (y*12)+m
    return res

def string_year_to_months(y):
    str1 =  "Convert {y} years to months.".format(y=y)
   # res = (y*365)+d
    return str1

def conver_year_to_months(y):
    res = (y*12)
    return res  

def string_convert_hr_and_min_into_min(h, m):
    str1 = "Convert the following into minutes: {h}hr {m} mins".format(h=h, m=m)
    return str1

def convert_hr_and_min_into_min(h, m):
    res = h*24 + m*24
    return res

def string_week_and_days_to_days(w, d):
   # str1 = "Convert 1 week 5 days 2 hours into hours."
    str1 = "Convert {w} week {d} days into days.".format(w=w, d=d)
    return str1

def convert_week_and_days_to_days(w, d):
    res = (w*7) + d
    return res
    

def format_string(first_no = None, second_no = None):
    str1 = "Round {first_no} to the nearest {second_no}.".format(first_no=first_no, second_no=second_no)
    return str1

def random_no_genrator(no=None):
    digits = set(range(1, 10))
    # We generate a random integer, 1 <= first <= 9
    first = random.randint(1, 9)
    last_remaening_digit = random.sample(digits - {first}, no)
    final_digit_no = str(first) + ''.join(map(str, last_remaening_digit))
    return final_digit_no


def round_no(n, d ):
    # Smaller multiple
    a = (n // d) * d
     
    # Larger multiple
    b = a + d
     
    # Return of closest of two
    return (b if n - a > b - n else a)


for i in range(0,16):
    id = 12 + i
    index = 9+i 
    indx = (index//2)
    
    questionId  =  "MA_4_01S0400" + str(id)   
    if i<5:
        four_digit_no = random_no_genrator(3)
        answer = round_no(int(four_digit_no),10)
        statement = format_string(four_digit_no, 10)
       # data = []
        data =pd.DataFrame({"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}, index=[indx])   
        #dic = {"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}   
        #data = pd.DataFrame(dic, index=[indx])
        df = df.append(data, ignore_index= False)
       # df.iloc[index] = [dic["questionId"], dic["topicId"], dic["subtopicId"], dic["Statement"], dic["Answer"], dic["Flags"], dic["icMCQ"], dic["Options"], dic["Image"] ]
        df = df.sort_index().reset_index(drop=True)
       # data=[dic]
       # print(data)
        #list(listdata[0].values()) = [questionId, "NAN", "NAN", statement, answer, "NAN", "NAN", "NAN", "NAN"]
        #df.iloc[index] = dic.values()
        #list(data[0].values())
        #df.sort_index().reset_index(drop=True)
        #index = index+1
    elif i>=5 and i<10:
        five_digit_no = random_no_genrator(4)
        answer = round_no(int(five_digit_no), 100)
        statement = format_string(five_digit_no, 100)
       # data = []
        data =pd.DataFrame({"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}, index=[indx])   
        #dic = {"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}   
      #  df.iloc[index] = [dic["questionId"], dic["topicId"], dic["subtopicId"], dic["Statement"], dic["Answer"], dic["Flags"], dic["icMCQ"], dic["Options"], dic["Image"] ]
        #final_lst.append(dic)
        #index = index+1
        #df.sort_index().reset_index(drop=True)
       # data = [dic]
        #data[0].values() = [questionId, "NAN", "NAN", statement, answer, "NAN", "NAN", "NAN", "NAN"]
        #df.iloc[index] =dic.values()
        #data[0].values()
        #data = pd.DataFrame(dic, index=[indx])
        df = df.append(data, ignore_index= False)
        df = df.sort_index().reset_index(drop=True)
    elif i>=10:
        six_digit_no = random_no_genrator(5)
        answer = round_no(int(six_digit_no), 1000)
        statement = format_string(six_digit_no, 1000)
        #data = []
        data =pd.DataFrame({"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}, index=[indx])   
       # data= [dic]
        #df.iloc[index] = [dic["questionId"], dic["topicId"], dic["subtopicId"], dic["Statement"], dic["Answer"], dic["Flags"], dic["icMCQ"], dic["Options"], dic["Image"] ]
       # data[0].values() = [questionId, "NAN", "NAN", statement, answer, "NAN", "NAN", "NAN", "NAN"]
        #df.iloc[index] = dic.values()
        #data[0].values()
        #data = pd.DataFrame(dic, index=[indx])
        df = df.append(data, ignore_index= False)
        df = df.sort_index().reset_index(drop=True)
#final_dic= {"round_qsn":final_lst}


#for i in range (0, len(lst)) :
 #   dic = lst[i]
  #  qsn_id = dic.get("questionId") 
   # j = 0
   # if qsn_id == "MA_4_01S040010" :
    #    j= i+1
    #lst.insert(j, final_lst)    
    
for i in range(0, 31):
    id = 12+i
    questionId = "MA_4_05S0400" +str(id)
    if i<5:
        y = random.randint(1, 3)
        d = random.randint(1, 365)
        statement = string_for_year_to_days(y, d)
        answer = conver_year_to_days(y, d)
        data =[{"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}]  
        df.loc[len(df.index)] = data[0]
        #df = df.append(data, ignore_index= False)
        df = df.sort_index().reset_index(drop=True)
        #df.append(data, ignore_index=True,sort=False)    
    elif i>=5 and i<10:
        y = random.randint(1, 20)
        d = random.randint(1, 11)
        statement = string_year_and_sum_months_to_months(y, d)
        answer = conver_year_sum_monts_to_months(y, d)
        data =[{"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}]  
        df.loc[len(df.index)] = data[0]
        df = df.sort_index().reset_index(drop=True)
    
    elif i>=15 and i<20:
        y = random.randint(1, 20)
        #d = random.randint(1, 11)
        statement = string_year_to_months(y)
        answer = conver_year_to_months(y)
        data =[{"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}]  
        df.loc[len(df.index)] = data[0]
        df = df.sort_index().reset_index(drop=True)
  
  
    
    elif i>=20 and i<25:
        h = random.randint(1, 23)
        m = random.randint(1, 60)
        statement = string_convert_hr_and_min_into_min(h, m)
        answer = convert_hr_and_min_into_min(h, m)
        data =[{"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}]  
        df.loc[len(df.index)] = data[0]
        df = df.sort_index().reset_index(drop=True)
    
    else:
        w = random.randint(2, 20)
        d = random.randint(1, 6)
        statement = string_week_and_days_to_days(w, d)
        answer = convert_week_and_days_to_days(w, d)
        data =[{"questionId": questionId, "topicId":"NAN", "subtopicId":"NAN", "Statement":statement, "Answer":answer, "Flags":"NAN", "icMCQ":"NAN", "Options":"NAN", "Image":"NAN"}]  
        df.loc[len(df.index)] = data[0]
        df = df.sort_index().reset_index(drop=True)

  
        
print(df)        

df.to_excel("D:/vt/Create-populate-excel-questions.xlsx", "Sheet1", index=False)